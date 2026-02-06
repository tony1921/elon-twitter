#!/usr/bin/env python3
"""
生成Excel表格 - Elon Musk推文数据
"""

import json
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def fetch_today_data():
    """获取今天的最新数据"""

    today = datetime.now().strftime("%Y-%m-%d")
    start_date = f"{today}T00:00:00.000Z"
    end_date = f"{today}T23:59:59.000Z"

    try:
        response = requests.get(
            "https://xtracker.polymarket.com/api/users/elonmusk/posts",
            params={'startDate': start_date, 'endDate': end_date},
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=30
        )

        if response.status_code == 200:
            data = response.json()
            if data.get('data'):
                return len(data['data'])
        return None
    except:
        return None


def load_historical_data():
    """加载历史数据"""

    with open('data/daily_tweets.json', 'r') as f:
        return json.load(f)


def create_excel():
    """创建Excel表格"""

    print("=" * 70)
    print("  📊 生成Excel表格")
    print("=" * 70)

    # 1. 更新今天的数据
    print("\n📡 获取今天的最新数据...")
    today_count = fetch_today_data()

    # 2. 加载历史数据
    print("📂 加载历史数据...")
    historical_data = load_historical_data()

    # 3. 更新今天的记录
    today_str = datetime.now().strftime("%Y-%m-%d")
    if today_count is not None:
        updated = False
        for record in historical_data:
            if record['date'] == today_str:
                record['count'] = today_count
                record['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break

        if not updated:
            historical_data.append({
                'date': today_str,
                'count': today_count,
                'source': 'xtracker_api',
                'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })

        print(f"  ✅ 今天({today_str}): {today_count} 条")

        # 保存更新后的数据
        with open('data/daily_tweets.json', 'w') as f:
            json.dump(historical_data, f, ensure_ascii=False, indent=2)
    else:
        print(f"  ℹ️  无法获取今天的实时数据")

    # 4. 创建DataFrame
    df = pd.DataFrame(historical_data)
    df = df[['date', 'count']]  # 只保留日期和数量

    # 按日期排序
    df = df.sort_values('date')

    # 5. 计算统计列
    df['weekday'] = pd.to_datetime(df['date']).dt.day_name()
    df['week_num'] = pd.to_datetime(df['date']).dt.isocalendar().week

    # 计算7天移动平均
    df['7day_avg'] = df['count'].rolling(window=7, min_periods=1).mean().round(1)

    # 计算与平均值的差异
    overall_avg = df['count'].mean()
    df['vs_avg'] = (df['count'] - overall_avg).round(1)

    # 6. 生成Excel文件
    excel_file = 'data/elon_musk_tweets.xlsx'

    print(f"\n📝 生成Excel文件: {excel_file}")

    # 创建Excel writer
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Sheet 1: 每日数据
        df.to_excel(writer, sheet_name='每日数据', index=False)

        # Sheet 2: 统计摘要
        # 创建统计数据
        stats_data = {
            '统计项': [
                '总天数',
                '总推文数',
                '平均每天',
                '最高单日',
                '最低单日',
                '中位数',
                '标准差',
                '数据范围',
                '最后更新'
            ],
            '值': [
                len(df),
                df['count'].sum(),
                f"{df['count'].mean():.1f}",
                df['count'].max(),
                df['count'].min(),
                f"{df['count'].median():.1f}",
                f"{df['count'].std():.1f}",
                f"{df['date'].min()} 至 {df['date'].max()}",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }

        # 添加周统计
        weekly_stats = df.groupby('week_num')['count'].agg(['sum', 'mean', 'max', 'min']).round(1)
        weekly_stats.columns = ['周总数', '周平均', '周最高', '周最低']

        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='统计摘要', index=False)

        weekly_stats.to_excel(writer, sheet_name='周统计')

    # 7. 美化Excel
    print("🎨 美化表格...")
    from openpyxl import load_workbook

    wb = load_workbook(excel_file)

    # 美化每日数据sheet
    ws1 = wb['每日数据']

    # 设置列宽
    ws1.column_dimensions['A'].width = 12  # 日期
    ws1.column_dimensions['B'].width = 10  # 数量
    ws1.column_dimensions['C'].width = 12  # 星期
    ws1.column_dimensions['D'].width = 10  # 周数
    ws1.column_dimensions['E'].width = 12  # 7日平均
    ws1.column_dimensions['F'].width = 12  # vs平均

    # 标题行样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 高亮最后几行（最近7天）
    last_rows = list(ws1.iter_rows(min_row=ws1.max_row - 6, max_row=ws1.max_row))
    recent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for row in last_rows:
        for cell in row:
            cell.fill = recent_fill

    # 美化统计摘要sheet
    ws2 = wb['统计摘要']
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 25

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 保存
    wb.save(excel_file)

    print(f"✅ Excel文件已生成: {excel_file}")

    # 8. 显示预览
    print(f"\n📊 数据预览（最近10天）:")
    print("=" * 90)
    print(df[['date', 'count', 'weekday', '7day_avg']].tail(10).to_string(index=False))
    print("=" * 90)

    print(f"\n📈 统计摘要:")
    print("=" * 90)
    print(f"  总天数: {len(df)} 天")
    print(f"  总推文: {df['count'].sum()} 条")
    print(f"  平均: {df['count'].mean():.1f} 条/天")
    print(f"  最高: {df['count'].max()} 条")
    print(f"  最低: {df['count'].min()} 条")
    print("=" * 90)

    return excel_file


if __name__ == "__main__":
    create_excel()
