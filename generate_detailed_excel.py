#!/usr/bin/env python3
"""
生成详细的推文数据Excel文件
"""

import json
import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 追踪期间
CURRENT_PERIOD = {
    'name': 'Feb 1 - Feb 28, 2026',
    'start': '2026-02-01T00:00:00.000Z',
    'end': '2026-02-28T23:59:59.000Z'
}

def fetch_all_posts():
    """获取所有推文数据"""
    url = 'https://xtracker.polymarket.com/api/users/elonmusk/posts'

    try:
        response = requests.get(
            url,
            params={
                'startDate': CURRENT_PERIOD['start'],
                'endDate': CURRENT_PERIOD['end']
            },
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=30
        )

        if response.status_code == 200:
            data = response.json()
            if data.get('data'):
                return data['data']

    except Exception as e:
        print(f"❌ 获取数据失败: {e}")

    return None

def analyze_posts(posts):
    """分析推文数据，使用不同的时区"""
    results = {
        'UTC': {},
        'EST': {},  # UTC-5
        'EDT': {},  # UTC-4 (夏令时)
        'detailed': []
    }

    for post in posts:
        created_at = post.get('createdAt', '')
        if not created_at:
            continue

        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))

        # UTC时间
        utc_date = dt.strftime('%Y-%m-%d')
        results['UTC'][utc_date] = results['UTC'].get(utc_date, 0) + 1

        # EST时间 (UTC-5)
        est_dt = dt - timedelta(hours=5)
        est_date = est_dt.strftime('%Y-%m-%d')
        results['EST'][est_date] = results['EST'].get(est_date, 0) + 1

        # EDT时间 (UTC-4)
        edt_dt = dt - timedelta(hours=4)
        edt_date = edt_dt.strftime('%Y-%m-%d')
        results['EDT'][edt_date] = results['EDT'].get(edt_date, 0) + 1

        # 详细信息
        results['detailed'].append({
            'UTC时间': created_at,
            'EST日期': est_date,
            'EDT日期': edt_date,
            '内容': post.get('text', ''),
            '链接': post.get('url', ''),
            '点赞': post.get('likes', 0),
            '转发': post.get('retweets', 0),
            '回复': post.get('replies', 0)
        })

    return results

def create_excel(results):
    """创建Excel文件"""

    wb = openpyxl.Workbook()

    # 删除默认sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 创建按日期统计的sheet
    ws_summary = wb.create_sheet('每日统计对比')
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 12

    # 标题
    headers = ['日期', 'UTC', 'EST (UTC-5)', 'EDT (UTC-4)']
    ws_summary.append(headers)

    # 标题样式
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 获取所有日期
    all_dates = set()
    all_dates.update(results['UTC'].keys())
    all_dates.update(results['EST'].keys())
    all_dates.update(results['EDT'].keys())

    for date in sorted(all_dates):
        ws_summary.append([
            date,
            results['UTC'].get(date, 0),
            results['EST'].get(date, 0),
            results['EDT'].get(date, 0)
        ])

    # 创建详细推文列表sheet
    ws_detail = wb.create_sheet('详细推文列表')
    ws_detail.column_dimensions['A'].width = 25
    ws_detail.column_dimensions['B'].width = 12
    ws_detail.column_dimensions['C'].width = 12
    ws_detail.column_dimensions['D'].width = 60
    ws_detail.column_dimensions['E'].width = 40
    ws_detail.column_dimensions['F'].width = 10
    ws_detail.column_dimensions['G'].width = 10
    ws_detail.column_dimensions['H'].width = 10

    # 标题
    detail_headers = ['UTC时间', 'EST日期', 'EDT日期', '内容', '链接', '点赞', '转发', '回复']
    ws_detail.append(detail_headers)

    # 标题样式
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据
    for post in results['detailed']:
        ws_detail.append([
            post['UTC时间'],
            post['EST日期'],
            post['EDT日期'],
            post['内容'][:200] if post['内容'] else '',  # 限制长度
            post['链接'],
            post['点赞'],
            post['转发'],
            post['回复']
        ])

    # 保存文件
    filename = f'data/elon_detailed_tweets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)

    print(f"✅ Excel文件已生成: {filename}")
    print(f"   - 总推文数: {len(results['detailed'])}")
    print(f"   - 统计天数: {len(all_dates)}")

    return filename

def main():
    print("=" * 70)
    print("  📊 生成详细推文数据Excel")
    print("=" * 70)

    # 获取数据
    print("\n⏳ 获取推文数据...")
    posts = fetch_all_posts()

    if not posts:
        print("❌ 未获取到数据")
        return

    print(f"✅ 获取到 {len(posts)} 条推文")

    # 分析数据
    print("\n⏳ 分析数据...")
    results = analyze_posts(posts)

    print("\n📊 时区对比:")
    print("-" * 70)
    print(f"{'日期':<12} {'UTC':<8} {'EST(UTC-5)':<12} {'EDT(UTC-4)':<12}")
    print("-" * 70)

    all_dates = set()
    all_dates.update(results['UTC'].keys())
    all_dates.update(results['EST'].keys())

    for date in sorted(all_dates):
        print(f"{date:<12} {results['UTC'].get(date, 0):<8} {results['EST'].get(date, 0):<12} {results['EDT'].get(date, 0):<12}")

    # 生成Excel
    print("\n⏳ 生成Excel文件...")
    filename = create_excel(results)

    print("\n" + "=" * 70)
    print("  ✅ 完成！")
    print("=" * 70)

if __name__ == "__main__":
    main()
